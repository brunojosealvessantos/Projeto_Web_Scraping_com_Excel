from selenium import webdriver
import openpyxl
import os
import time


class BuscaPrecos:
    def __init__(self):
        self.driver = webdriver.Chrome(
            executable_path=os.getcwd()+os.sep+'chromedriver.exe')

    def iniciar(self):
        self.numero_pagina_atual = 1
        self.driver.get(
            f'https://ba.olx.com.br/autos-e-pecas/caminhoes?o={self.numero_pagina_atual}')
        self.criar_planilha()
        self.encontrar_valores_na_pagina()

    def criar_planilha(self):
        self.planilha = openpyxl.Workbook()
        self.planilha.create_sheet('valores')
        self.planilha_valores = self.planilha['valores']
        self.planilha_valores.cell(row=1, column=1, value='Titulo')
        self.planilha_valores.cell(row=1, column=2, value='Localização')
        self.planilha_valores.cell(row=1, column=3, value='Preço')

    def encontrar_valores_na_pagina(self):
        try:
            while True:
                self.titulos = self.driver.find_elements_by_xpath(
                    "//h2[@class='fnmrjs-10 deEIZJ']")
                self.localizacao = self.driver.find_elements_by_xpath(
                    "//p[@class='fnmrjs-16 jqSHIm']")
                self.preco = self.driver.find_elements_by_xpath(
                    "//p[@class='fnmrjs-13 hdwqVC']")
                self.armazenar_valores_na_planilha()
                self.navegar_para_proxima_pagina()
        except Exception as erro:
            print('Não há mais pesquisas para extrair informações')
            print(erro)

    def navegar_para_proxima_pagina(self):
        self.numero_pagina_atual += 1
        self.driver.get(
            f'https://ba.olx.com.br/autos-e-pecas/caminhoes?o={self.numero_pagina_atual}')

    def armazenar_valores_na_planilha(self):
        for indice in range(0, len(self.titulos)):
            nova_linha = [self.titulos[indice].text,
                          self.localizacao[indice].text, self.preco[indice].text]
            self.planilha_valores.append(nova_linha)
        self.planilha.save('Planilha de Preços')


bot = BuscaPrecos()
bot.iniciar()
